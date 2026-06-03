"""collate_pos.py — Collate all purchase orders modified this year into one Excel.

Pipeline (mirrors watch_orders.py's approach):
  1. Find every PO PDF modified on/after 2026-01-01 (ignores the duplicate .xls copies).
  2. pdfplumber extracts the raw text.
  3. Claude Haiku parses each PO's messy text into structured line items.
  4. Lines are deduped per (supplier, description), keeping the most recent unit cost.
  5. Writes scripts/po_collation.xlsx:
        - "Purchases"  : every distinct item, grouped by supplier, with a Category dropdown.
        - "Suppliers"  : contact/address details pulled from the PO headers.
        - "Categories" : the app's material categories (the dropdown source).

Run:  python scripts/collate_pos.py            (full run)
      python scripts/collate_pos.py --limit 10  (test on first 10 POs)
"""
import sys, os, re, json, time, argparse
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock

sys.path.insert(0, str(Path(__file__).parent.parent))
import pdfplumber
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname

from app import create_app
from models import Category, Supplier

PO_DIR = Path(r"C:\Users\conta\DT Solutions LTD\DTCompany - Documents\Company\DTS Admin\Purchase Orders")
CUT = datetime(2026, 1, 1)
OUT_PATH = Path(__file__).parent / "po_collation.xlsx"
CACHE_PATH = Path(__file__).parent / ".po_cache.json"
MODEL = "claude-haiku-4-5-20251001"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # yellow = fill this in
GROUP_FILL = PatternFill("solid", fgColor="DCE6F1")    # supplier separator
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

client = anthropic.Anthropic(max_retries=8)   # SDK backs off (honours Retry-After) on 429

_cache_lock = Lock()

SYS = ("You extract line items from a UK purchase order PDF. DT Solutions is the BUYER; "
       "the named supplier is who they buy FROM. The text is messy: the pound sign often "
       "appears as the replacement char, and digits in prices have stray spaces "
       "(e.g. '1 19.25' means 119.25, '1 ,024.80' means 1024.80). Return ONLY valid JSON.")

PROMPT = """Filename: {name}

RAW PO TEXT:
{text}

Return ONLY this JSON object (no prose, no markdown fences):
{{
  "supplier_name": "the supplier bought FROM (the TO: block); else infer from filename",
  "supplier_address": "single string, comma-separated, or null",
  "supplier_phone": "or null",
  "supplier_email": "or null",
  "supplier_contact": "person name if shown, or null",
  "po_ref": "the P/O NO. value",
  "po_date": "YYYY-MM-DD from the DATE field, or null",
  "lines": [
    {{
      "description": "the material/product description, cleaned",
      "unit": "sheet/box/roll/pcs/etc or null",
      "qty": number or null,
      "unit_price": number or null,
      "nominal_code": "4-digit QuickBooks nominal code if present, else null",
      "ref_note": "any short 'what it's for' note (e.g. 'Hinge Jig','TCP'), else null"
    }}
  ]
}}
Rules:
- Include EVERY distinct material line that has a description, even if qty/amount is blank.
- unit_price is the per-UNIT price column, NOT the line total/amount.
- Exclude carriage, subtotal, VAT and total lines, and pure notes like 'I can collect when ready'.
- If a price shows stray spaces, collapse them ('1 19.25' -> 119.25).
- IMPORTANT: the supplier_* fields must be the SUPPLIER's details, never DT Solutions' own.
  Ignore DT's buyer/delivery details: 'DT Solutions', 'Unit 7 Woodrow Way', 'Unit 4 Tuffley
  Industrial Park', 'Pearce Way', 'Gloucester', 'GL2 5YD', 'GL2 5DX', and phone '01452 332727'.
  If the only phone/address on the page is DT's, return null for that supplier field."""


def in_scope_pdfs(limit=None):
    pdfs = [p for p in PO_DIR.rglob("*.pdf")
            if datetime.fromtimestamp(p.stat().st_mtime) >= CUT]
    pdfs.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return pdfs[:limit] if limit else pdfs


def extract_text(path):
    try:
        with pdfplumber.open(path) as pdf:
            txt = "\n".join((pg.extract_text() or "") for pg in pdf.pages)
        return txt.replace("�", "£")
    except Exception as e:
        return f"__ERROR__ {e}"


def parse_po(path):
    mtime = datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d")
    text = extract_text(path)
    if text.startswith("__ERROR__"):
        return {"_file": path.name, "_error": text}
    last = None
    for attempt in range(6):
        try:
            resp = client.messages.create(
                model=MODEL, max_tokens=3000, temperature=0,
                system=SYS,
                messages=[{"role": "user",
                           "content": PROMPT.format(name=path.name, text=text[:8000])}],
            )
            raw = resp.content[0].text.strip()
            m = re.search(r"\{.*\}", raw, re.DOTALL)
            data = json.loads(m.group(0) if m else raw)
            data["_file"] = path.name
            data["_mtime"] = mtime
            return data
        except anthropic.RateLimitError as e:
            last = str(e)
            time.sleep(min(2 ** attempt, 30))   # extra backoff beyond the SDK's own
        except Exception as e:
            last = str(e)
            time.sleep(1.5)
    return {"_file": path.name, "_error": last}


def norm_supplier(name):
    if not name:
        return "unknown"
    s = re.sub(r"[^a-z0-9 ]", "", name.lower())
    s = re.sub(r"\b(ltd|limited|services|service|uk|co|company|the|plc)\b", "", s)
    return re.sub(r"\s+", " ", s).strip() or "unknown"


def norm_desc(d):
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]", " ", (d or "").lower())).strip()


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=None)
    args = ap.parse_args()

    app = create_app()
    with app.app_context():
        categories = [c.name for c in Category.query.order_by(Category.name).all()]
        db_suppliers = [s.name for s in Supplier.query.all()]

    pdfs = in_scope_pdfs(args.limit)

    # Cache (resumable): key = "filename|mtime" -> parsed result. Only successes are cached.
    cache = {}
    if CACHE_PATH.exists():
        cache = json.loads(CACHE_PATH.read_text(encoding="utf-8"))

    def ckey(p):
        return f"{p.name}|{datetime.fromtimestamp(p.stat().st_mtime).strftime('%Y-%m-%d')}"

    todo = [p for p in pdfs if ckey(p) not in cache]
    print(f"{len(pdfs)} POs in scope; {len(pdfs) - len(todo)} cached, parsing {len(todo)} "
          f"with {MODEL} (4 workers)...")

    results, errors = [], []
    done = 0

    def save_cache():
        with _cache_lock:
            CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf-8")

    with ThreadPoolExecutor(max_workers=4) as ex:
        futs = {ex.submit(parse_po, p): p for p in todo}
        for fut in as_completed(futs):
            r = fut.result()
            done += 1
            if "_error" not in r:
                with _cache_lock:
                    cache[ckey(futs[fut])] = r
            else:
                errors.append(r)
            if done % 10 == 0 or done == len(todo):
                print(f"  {done}/{len(todo)} parsed ({len(errors)} errors)")
                save_cache()
    save_cache()

    # Pull every in-scope PO's result from the cache (cached + freshly parsed).
    results = [cache[ckey(p)] for p in pdfs if ckey(p) in cache]

    # ---- Collate suppliers + deduped line items ----
    suppliers = {}   # key -> {display, address, phone, email, contact}
    items = {}       # (supkey, desc_norm) -> row dict
    for r in results:
        skey = norm_supplier(r.get("supplier_name"))
        sup = suppliers.setdefault(skey, {"display": r.get("supplier_name") or skey,
                                          "address": None, "phone": None,
                                          "email": None, "contact": None})
        for f in ("address", "phone", "email", "contact"):
            if not sup[f] and r.get(f"supplier_{f}"):
                sup[f] = r[f"supplier_{f}"]
        pdate = r.get("po_date") or r.get("_mtime")
        for ln in r.get("lines", []):
            desc = (ln.get("description") or "").strip()
            if not desc:
                continue
            key = (skey, norm_desc(desc))
            cur = items.get(key)
            row = {"supplier": sup["display"], "supkey": skey, "description": desc,
                   "unit": ln.get("unit"), "unit_price": ln.get("unit_price"),
                   "nominal": ln.get("nominal_code"), "ref": ln.get("ref_note"),
                   "last_po": r.get("po_ref"), "last_date": pdate, "seen": 1}
            if not cur:
                items[key] = row
            else:
                cur["seen"] += 1
                # keep the most recent record's price/description
                if (pdate or "") >= (cur["last_date"] or ""):
                    cur["last_date"], cur["last_po"] = pdate, r.get("po_ref")
                    cur["description"] = desc
                    if ln.get("unit_price") is not None:
                        cur["unit_price"] = ln.get("unit_price")
                    if ln.get("unit"):
                        cur["unit"] = ln.get("unit")
                    if ln.get("nominal_code"):
                        cur["nominal"] = ln.get("nominal_code")

    rows = sorted(items.values(), key=lambda r: (r["supplier"].lower(), r["description"].lower()))
    print(f"\nCollated {len(rows)} distinct items across {len(suppliers)} suppliers.")

    # ---- Build workbook ----
    wb = Workbook()

    # Categories reference (dropdown source) + IGNORE option
    ws_cat = wb.active
    ws_cat.title = "Categories"
    ws_cat.append(["Category (dropdown options)"])
    cat_options = categories + ["IGNORE - not a material"]
    for c in cat_options:
        ws_cat.append([c])
    style_header(ws_cat, 1)
    ws_cat.column_dimensions["A"].width = 32
    wb.defined_names.add(DefinedName(
        "cat_options", attr_text=f"{quote_sheetname('Categories')}!$A$2:$A${len(cat_options)+1}"))

    # Purchases
    ws = wb.create_sheet("Purchases", 0)
    headers = ["Supplier", "Category", "Description", "Unit", "Unit Cost (£)",
               "QB Nominal", "Ref / Note", "Times Ordered", "Last PO", "Last PO Date"]
    ws.append(headers)
    style_header(ws, len(headers))
    widths = [26, 26, 60, 8, 12, 10, 18, 12, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    prev = None
    for r in rows:
        if prev is not None and r["supkey"] != prev:
            ws.append([])  # blank separator row between suppliers
        ws.append([r["supplier"], None, r["description"], r["unit"], r["unit_price"],
                   r["nominal"], r["ref"], r["seen"], r["last_po"], r["last_date"]])
        cell = ws.cell(ws.max_row, 5)
        if isinstance(r["unit_price"], (int, float)):
            cell.number_format = "£#,##0.00"
        ws.cell(ws.max_row, 2).fill = INPUT_FILL   # Category cell = yellow
        prev = r["supkey"]

    dv = DataValidation(type="list", formula1="=cat_options", allow_blank=True)
    dv.error = "Pick a category from the dropdown (see Categories tab)."
    dv.errorTitle = "Choose a category"
    ws.add_data_validation(dv)
    dv.add(f"B2:B{ws.max_row + 5}")

    # Suppliers
    ws_sup = wb.create_sheet("Suppliers")
    sup_headers = ["Supplier", "Contact", "Email", "Phone", "Address", "Already in app?"]
    ws_sup.append(sup_headers)
    style_header(ws_sup, len(sup_headers))
    for i, w in enumerate([26, 22, 30, 22, 50, 14], 1):
        ws_sup.column_dimensions[ws_sup.cell(1, i).column_letter].width = w
    def match_db(po_norm):
        po_tok = set(po_norm.split())
        for name in db_suppliers:
            dn = norm_supplier(name)
            if not dn:
                continue
            if po_norm == dn or po_norm in dn or dn in po_norm:
                return name
            dtok = set(dn.split())
            if po_tok & dtok and po_norm.split()[:1] == dn.split()[:1]:
                return name
        return None

    for skey, s in sorted(suppliers.items(), key=lambda kv: kv[1]["display"].lower()):
        m = match_db(skey)
        in_db = f"yes — {m}" if m else "NEW"
        ws_sup.append([s["display"], s["contact"], s["email"], s["phone"], s["address"], in_db])

    wb.save(OUT_PATH)
    print(f"\nSaved: {OUT_PATH}")
    print(f"  Purchases : {len(rows)} distinct items")
    print(f"  Suppliers : {len(suppliers)}")
    print(f"  Categories: {len(categories)} (dropdown)")
    if errors:
        print(f"\n{len(errors)} file(s) failed to parse (need manual look):")
        for e in errors:
            print(f"  - {e['_file']}: {e['_error'][:80]}")


if __name__ == "__main__":
    main()
