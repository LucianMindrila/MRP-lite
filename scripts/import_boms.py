"""import_boms.py — Load filled-in BOM rows from bom_template.xlsx back into the DB.

Reads the "Enter BOMs" tab, validates every row, and creates/updates BOM items.

    python scripts/import_boms.py                 # validate + import
    python scripts/import_boms.py --dry-run        # validate + preview only, no writes
    python scripts/import_boms.py --file other.xlsx

Safety:
    - Strict validation FIRST. If any row has an unknown product/material code or a
      bad quantity, NOTHING is written — it prints the problem rows so you can fix
      them and re-run. (A typo can never half-import.)
    - Rows where Material Code and Qty are both blank are skipped (products you
      haven't filled in yet).
    - Re-runnable: a (product, material) pair that already exists has its quantity
      updated rather than duplicated.

After a successful import, refresh the snapshot:
    python scripts/export_db.py
"""
import sys
import argparse
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from openpyxl import load_workbook

from app import create_app
from models import db, Product, Material, BOMItem

DEFAULT_FILE = Path(__file__).parent / 'bom_template.xlsx'
SHEET = 'Enter BOMs'


def cell_str(v):
    """Normalise a cell value to a trimmed string ('' for blank)."""
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def parse_rows(path):
    wb = load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"ERROR: sheet '{SHEET}' not found in {path}")
    ws = wb[SHEET]

    rows = []
    for r in range(2, ws.max_row + 1):
        prod_code = cell_str(ws.cell(r, 1).value)   # A
        mat_code = cell_str(ws.cell(r, 3).value)     # C
        qty_raw = ws.cell(r, 5).value                # E
        # Untouched / not-yet-filled row: no material and no qty -> skip silently.
        if not mat_code and (qty_raw is None or cell_str(qty_raw) == ''):
            continue
        rows.append({'row': r, 'product': prod_code, 'material': mat_code, 'qty': qty_raw})
    return rows


def validate(rows, app):
    """Return (valid_entries, errors). valid_entries hold resolved IDs."""
    errors = []
    valid = []
    with app.app_context():
        prod_by_code = {p.code: p for p in Product.query.all()}
        mat_by_code = {m.code: m for m in Material.query.all()}
        for row in rows:
            r = row['row']
            p = prod_by_code.get(row['product'])
            m = mat_by_code.get(row['material'])
            if not row['product']:
                errors.append(f"Row {r}: product code is blank")
                continue
            if not p:
                errors.append(f"Row {r}: unknown product code '{row['product']}'")
            if not row['material']:
                errors.append(f"Row {r}: material code is blank")
            elif not m:
                errors.append(f"Row {r}: unknown material code '{row['material']}'")
            try:
                qty = float(row['qty'])
            except (TypeError, ValueError):
                errors.append(f"Row {r}: quantity '{row['qty']}' is not a number")
                qty = None
            else:
                if qty <= 0:
                    errors.append(f"Row {r}: quantity must be greater than 0 (got {qty})")
            if p and m and qty and qty > 0:
                valid.append({'row': r, 'product_id': p.id, 'material_id': m.id,
                              'product': p.code, 'material': m.code, 'qty': qty})

    # Flag the same material listed twice for one product within the sheet.
    seen = {}
    for v in valid:
        key = (v['product_id'], v['material_id'])
        if key in seen:
            errors.append(f"Row {v['row']}: material '{v['material']}' listed twice for "
                          f"product '{v['product']}' (also row {seen[key]})")
        else:
            seen[key] = v['row']
    return valid, errors


def apply_entries(valid, app, dry_run):
    created = updated = 0
    with app.app_context():
        for v in valid:
            existing = BOMItem.query.filter_by(
                product_id=v['product_id'], material_id=v['material_id']).first()
            if existing:
                if existing.qty_per_unit != v['qty']:
                    if not dry_run:
                        existing.qty_per_unit = v['qty']
                    updated += 1
            else:
                if not dry_run:
                    db.session.add(BOMItem(product_id=v['product_id'],
                                           material_id=v['material_id'],
                                           qty_per_unit=v['qty']))
                created += 1
        if not dry_run:
            db.session.commit()
    return created, updated


def main():
    parser = argparse.ArgumentParser(description='Import BOMs from the Excel template')
    parser.add_argument('--file', default=str(DEFAULT_FILE), help='Path to the filled template')
    parser.add_argument('--dry-run', action='store_true', help='Validate + preview, write nothing')
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        sys.exit(f"ERROR: {path} not found. Run `python scripts/make_bom_template.py` first.")

    app = create_app()
    rows = parse_rows(path)
    if not rows:
        sys.exit("No filled-in BOM rows found. Nothing to import.")

    valid, errors = validate(rows, app)

    if errors:
        print(f"FOUND {len(errors)} PROBLEM(S) - nothing was written. Fix these and re-run:\n")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)

    created, updated = apply_entries(valid, app, args.dry_run)

    tag = "[DRY RUN] would " if args.dry_run else ""
    print(f"{len(rows)} filled rows read, {len(valid)} valid BOM lines.")
    print(f"{tag}create {created} new BOM line(s), update {updated} existing.")
    # How many products are now fully covered
    products_touched = {v['product_id'] for v in valid}
    print(f"BOM lines span {len(products_touched)} product(s).")
    if not args.dry_run:
        print("\nDone. Now refresh the snapshot:  python scripts/export_db.py")


if __name__ == '__main__':
    main()
