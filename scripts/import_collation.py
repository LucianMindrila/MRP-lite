"""import_collation.py — Import suppliers and materials from po_collation.xlsx into the DB.

Reads the "Purchases" and "Suppliers" sheets, then:
  1. Creates or updates Suppliers (fills in missing contact/email/phone/address only).
  2. Creates Materials, skipping IGNORE rows and anything already in the DB
     (matched by normalised name + supplier).

Run:
  python scripts/import_collation.py            # preview (dry-run by default)
  python scripts/import_collation.py --commit   # actually write to the database
"""
import sys, io, re, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook
from app import create_app
from models import db, Supplier, Material, Category

XL = Path(__file__).parent / "po_collation.xlsx"
IGNORE_CAT = "IGNORE - not a material"


# ── helpers ──────────────────────────────────────────────────────────────────

def norm(s):
    """Lowercase, strip punctuation, collapse whitespace — for fuzzy matching."""
    s = re.sub(r"[^a-z0-9 ]", "", (s or "").lower())
    s = re.sub(r"\b(ltd|limited|services|service|uk|co|company|the|plc)\b", "", s)
    return re.sub(r"\s+", " ", s).strip()


def make_code(description, existing_codes):
    """Generate a short unique material code from a description."""
    slug = re.sub(r"[^A-Z0-9 ]", "", description.upper())
    slug = re.sub(r"\s+", "-", slug.strip())[:28].rstrip("-")
    code = slug
    n = 2
    while code in existing_codes:
        code = f"{slug[:25]}-{n}"
        n += 1
    existing_codes.add(code)
    return code


def norm_unit(u):
    """Return a clean unit string, defaulting to 'pcs'."""
    if not u:
        return "pcs"
    u = str(u).strip().lower()
    return u if u else "pcs"


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--commit", action="store_true",
                    help="Write to the database (default is dry-run / preview only)")
    args = ap.parse_args()
    dry = not args.commit

    wb = load_workbook(XL, data_only=True)

    # ── 1. Read Suppliers sheet ───────────────────────────────────────────────
    ws_sup = wb["Suppliers"]
    xl_suppliers = {}   # norm_name -> raw dict
    for row in ws_sup.iter_rows(min_row=2, values_only=True):
        name = (row[0] or "").strip()
        if not name:
            continue
        xl_suppliers[norm(name)] = {
            "display": name,
            "contact": (row[1] or "").strip() or None,
            "email":   (row[2] or "").strip() or None,
            "phone":   (row[3] or "").strip() or None,
            "address": (row[4] or "").strip() or None,
        }

    # ── 2. Read Purchases sheet ───────────────────────────────────────────────
    ws_pur = wb["Purchases"]
    # cols: Supplier(0) Category(1) Description(2) Unit(3) UnitCost(4) ...
    xl_rows = []
    for row in ws_pur.iter_rows(min_row=2, values_only=True):
        sup_name = (row[0] or "").strip()
        category = (row[1] or "").strip()
        desc     = (row[2] or "").strip()
        unit     = row[3]
        price    = row[4]
        if not desc:
            continue   # blank separator row
        if category == IGNORE_CAT or not category:
            continue   # skip non-materials and uncategorised
        xl_rows.append({
            "supplier": sup_name,
            "category": category,
            "description": desc,
            "unit": unit,
            "unit_price": price,
        })

    app = create_app()
    with app.app_context():

        # ── 3. Resolve / create suppliers ─────────────────────────────────────
        db_suppliers = {norm(s.name): s for s in Supplier.query.all()}
        sup_created = sup_updated = 0
        sup_map = {}   # norm_name -> Supplier ORM object

        for nname, xl in xl_suppliers.items():
            existing = db_suppliers.get(nname)
            if existing:
                changed = False
                for field in ("contact", "email", "phone", "address"):
                    if not getattr(existing, field) and xl[field]:
                        if not dry:
                            setattr(existing, field, xl[field])
                        changed = True
                if changed:
                    sup_updated += 1
                sup_map[nname] = existing
            else:
                s = Supplier(
                    name=xl["display"],
                    contact=xl["contact"],
                    email=xl["email"],
                    phone=xl["phone"],
                    address=xl["address"],
                )
                if not dry:
                    db.session.add(s)
                    db.session.flush()   # get the id
                sup_map[nname] = s
                sup_created += 1

        # ── 4. Resolve categories ──────────────────────────────────────────────
        cat_map = {c.name: c for c in Category.query.all()}

        # ── 5. Resolve / create materials ─────────────────────────────────────
        existing_codes = {m.code for m in Material.query.all()}
        # Key for dedup: (supplier_id_or_None, norm_description)
        existing_mats = {}
        for m in Material.query.all():
            existing_mats[(m.supplier_id, norm(m.name))] = m

        mat_created = mat_skipped = mat_updated = 0
        preview_create = []
        preview_skip = []
        preview_update = []

        for row in xl_rows:
            cat = cat_map.get(row["category"])
            if not cat:
                print(f"  WARN: category not found in DB: {row['category']!r} — skipping row")
                continue

            sup_norm = norm(row["supplier"])
            sup_obj = sup_map.get(sup_norm)
            sup_id = sup_obj.id if (sup_obj and hasattr(sup_obj, "id") and sup_obj.id) else None

            dedup_key = (sup_id, norm(row["description"]))
            existing = existing_mats.get(dedup_key)

            if existing:
                # Material already in DB — update price if we now have one and it was 0
                price = row["unit_price"]
                if isinstance(price, (int, float)) and price > 0 and existing.cost_price == 0:
                    if not dry:
                        existing.cost_price = float(price)
                    mat_updated += 1
                    preview_update.append(
                        f"  UPDATE price: {existing.code}  {existing.name[:40]}  -> £{price:.2f}")
                else:
                    mat_skipped += 1
                    preview_skip.append(
                        f"  SKIP (exists): {existing.code}  {existing.name[:40]}")
            else:
                code = make_code(row["description"], existing_codes)
                price = row["unit_price"]
                mat = Material(
                    code=code,
                    name=row["description"],
                    unit=norm_unit(row["unit"]),
                    cost_price=float(price) if isinstance(price, (int, float)) else 0.0,
                    supplier_id=sup_id,
                    category_id=cat.id,
                )
                if not dry:
                    db.session.add(mat)
                mat_created += 1
                sup_label = sup_obj.name if sup_obj else "—"
                preview_create.append(
                    f"  CREATE: {code:30s}  {row['description'][:40]:42s}  [{row['category']}]  {sup_label}")
                existing_mats[dedup_key] = mat   # prevent double-create within this run

        if not dry:
            db.session.commit()

    # ── 6. Report ─────────────────────────────────────────────────────────────
    tag = "[DRY RUN] " if dry else ""

    print(f"\n{'='*60}")
    print(f"{tag}SUPPLIERS")
    print(f"  {sup_created} new  |  {sup_updated} updated (filled-in blanks)  |  "
          f"{len(xl_suppliers) - sup_created - sup_updated} already complete")

    print(f"\n{tag}MATERIALS")
    print(f"  {mat_created} to create  |  {mat_updated} price updates  |  {mat_skipped} already exist")

    if preview_create:
        print(f"\n  --- Creating ({len(preview_create)}) ---")
        for line in preview_create:
            print(line)

    if preview_update:
        print(f"\n  --- Updating price ({len(preview_update)}) ---")
        for line in preview_update:
            print(line)

    if preview_skip and dry:
        print(f"\n  --- Skipping / already in DB ({len(preview_skip)}) ---")
        for line in preview_skip[:20]:
            print(line)
        if len(preview_skip) > 20:
            print(f"  ... and {len(preview_skip) - 20} more")

    if dry:
        print(f"\n{'='*60}")
        print("This was a DRY RUN — nothing was written.")
        print("Re-run with --commit to apply.\n")
    else:
        print(f"\nDone. Database updated.")
        print("Next: python scripts/export_db.py\n")


if __name__ == "__main__":
    main()
