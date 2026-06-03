# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from openpyxl import load_workbook
from collections import Counter

wb = load_workbook("scripts/po_collation.xlsx", data_only=True)
ws = wb["Purchases"]
rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(r)]
print("Data rows (non-blank):", len(rows))

cat = Counter((r[1] or "(blank)") for r in rows)
print("\nCategory assignment counts:")
for k, v in sorted(cat.items(), key=lambda x: -x[1]):
    print(f"  {v:4}  {k}")

blank = [r for r in rows if not r[1]]
print(f"\nRows still without a category: {len(blank)}")
for r in blank[:20]:
    desc = (r[2] or "")[:50]
    print(f"   - {r[0]} | {desc}")
