"""make_bom_template.py — Generate an Excel template for filling in missing BOMs.

Creates scripts/bom_template.xlsx with:
  - "Enter BOMs"   : one starter row per product that currently has no BOM.
                     Material Code is a dropdown (valid codes only); Material Name
                     auto-fills via lookup so you can confirm you picked the right one.
                     Add extra rows for products that need more than one material.
  - "Materials"    : reference list of all materials (the dropdown source).
  - "Products"     : reference list of the products needing a BOM.
  - "Instructions" : how to fill it in.

Workflow:
    python scripts/make_bom_template.py     # create the template
    ... fill it in manually in Excel ...
    python scripts/import_boms.py            # load the filled rows back into the DB

Run from the project root with the venv active.
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname

from app import create_app
from models import Material, Product

OUT_PATH = Path(__file__).parent / 'bom_template.xlsx'
MAX_ROWS = 600  # how far down to extend dropdowns / lookup formulas for added rows

HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(bold=True, color='FFFFFF')
INPUT_FILL = PatternFill('solid', fgColor='FFF2CC')   # pale yellow = "type here"
AUTO_FILL = PatternFill('solid', fgColor='EDEDED')     # grey = auto / don't touch
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'


def build():
    app = create_app()
    with app.app_context():
        materials = Material.query.order_by(Material.code).all()
        no_bom = [p for p in Product.query.order_by(Product.code).all() if not p.bom_items]
        mats = [(m.code, m.name, m.unit, m.supplier.name if m.supplier else '') for m in materials]
        prods = [(p.code, p.name) for p in no_bom]

    wb = Workbook()

    # ---- Reference sheet: Materials ----
    ws_mat = wb.active
    ws_mat.title = 'Materials'
    ws_mat.append(['Code', 'Name', 'Unit', 'Supplier'])
    for row in mats:
        ws_mat.append(list(row))
    style_header(ws_mat, 4)
    for col, w in zip('ABCD', (22, 55, 10, 28)):
        ws_mat.column_dimensions[col].width = w

    # ---- Reference sheet: Products (those needing a BOM) ----
    ws_prod = wb.create_sheet('Products')
    ws_prod.append(['Code', 'Name'])
    for row in prods:
        ws_prod.append(list(row))
    style_header(ws_prod, 2)
    ws_prod.column_dimensions['A'].width = 22
    ws_prod.column_dimensions['B'].width = 60

    # ---- Named ranges for the dropdowns ----
    mat_last = len(mats) + 1
    prod_last = len(prods) + 1
    wb.defined_names.add(DefinedName(
        'material_codes', attr_text=f"{quote_sheetname('Materials')}!$A$2:$A${mat_last}"))
    wb.defined_names.add(DefinedName(
        'product_codes', attr_text=f"{quote_sheetname('Products')}!$A$2:$A${prod_last}"))

    # ---- Entry sheet ----
    ws = wb.create_sheet('Enter BOMs', 0)  # make it the first/active tab
    ws.append(['Product Code', 'Product Name', 'Material Code', 'Material Name', 'Qty per Unit'])
    style_header(ws, 5)
    for col, w in zip('ABCDE', (22, 45, 22, 45, 14)):
        ws.column_dimensions[col].width = w

    # Pre-fill one starter row per product needing a BOM.
    for i, (code, _name) in enumerate(prods):
        ws.cell(row=2 + i, column=1, value=code)

    # Formulas (name lookups) + styling for every potential data row.
    for r in range(2, MAX_ROWS + 1):
        ws.cell(row=r, column=2).value = (
            f'=IFERROR(VLOOKUP(A{r},Products!$A:$B,2,FALSE),"")')
        ws.cell(row=r, column=4).value = (
            f'=IFERROR(VLOOKUP(C{r},Materials!$A:$B,2,FALSE),"")')
        ws.cell(row=r, column=1).fill = INPUT_FILL
        ws.cell(row=r, column=3).fill = INPUT_FILL
        ws.cell(row=r, column=5).fill = INPUT_FILL
        ws.cell(row=r, column=2).fill = AUTO_FILL
        ws.cell(row=r, column=4).fill = AUTO_FILL

    # Dropdowns
    dv_prod = DataValidation(type='list', formula1='=product_codes', allow_blank=True)
    dv_prod.error = 'Pick a product code from the list (see Products tab).'
    dv_prod.errorTitle = 'Unknown product code'
    dv_mat = DataValidation(type='list', formula1='=material_codes', allow_blank=True)
    dv_mat.error = 'Pick a material code from the list (see Materials tab).'
    dv_mat.errorTitle = 'Unknown material code'
    dv_qty = DataValidation(type='decimal', operator='greaterThan', formula1='0', allow_blank=True)
    dv_qty.error = 'Quantity must be a number greater than 0.'
    dv_qty.errorTitle = 'Invalid quantity'
    ws.add_data_validation(dv_prod)
    ws.add_data_validation(dv_mat)
    ws.add_data_validation(dv_qty)
    dv_prod.add(f'A2:A{MAX_ROWS}')
    dv_mat.add(f'C2:C{MAX_ROWS}')
    dv_qty.add(f'E2:E{MAX_ROWS}')

    # ---- Instructions ----
    ws_help = wb.create_sheet('Instructions')
    lines = [
        ('How to fill in this BOM template', True),
        ('', False),
        ('A BOM (Bill of Materials) lists which materials go into making one unit of a product,', False),
        ('and how many of each. Fill in the "Enter BOMs" tab.', False),
        ('', False),
        ('1. The yellow columns are for you to fill: Material Code and Qty per Unit.', False),
        ('2. The 37 products that currently have no BOM are already listed in column A.', False),
        ('3. For each product, pick a Material Code from the dropdown. The Material Name', False),
        ('   fills in automatically (grey columns) so you can check it is the right one.', False),
        ('4. Enter Qty per Unit = how many of that material are needed to make ONE product.', False),
        ('5. If a product needs MORE THAN ONE material, add a new row underneath and put', False),
        ('   the SAME product code again, then the next material + qty. Repeat as needed.', False),
        ('6. Leave a product blank if you do not know its BOM yet — you can come back to it.', False),
        ('', False),
        ('When done, save the file (keep the name bom_template.xlsx) and tell Claude, or run:', False),
        ('    python scripts/import_boms.py', False),
        ('', False),
        ('Tip: the "Materials" tab lists every material code, name, unit and supplier for reference.', False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws_help.cell(row=i, column=1, value=text)
        if bold:
            c.font = Font(bold=True, size=14)
    ws_help.column_dimensions['A'].width = 100

    wb.save(OUT_PATH)
    print(f"Template written to: {OUT_PATH}")
    print(f"  Products needing a BOM : {len(prods)} (pre-listed in 'Enter BOMs')")
    print(f"  Materials to choose from: {len(mats)} (dropdown + 'Materials' tab)")


if __name__ == '__main__':
    build()
